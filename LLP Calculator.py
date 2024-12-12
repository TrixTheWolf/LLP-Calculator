#Author: Ben Renner
#Title: LLP Calculator
import openpyxl

dataframe = openpyxl.load_workbook("V2500_LLP_PRICING_CALCULATOR.xlsx", data_only=True) # Define variable to load the dataframe
dataframe1 = dataframe.active # Define variable to read sheet
print('Welcome to the LLP Pricing Calcuator')

#loop lasts until user wants to end
while True:
#User Inputs
    part_nb = input("Enter the Part Number: ").upper()
    cycle_rm = int(input("Enter the Total Cycles Remaining(numbers only): "))
    matched_row = None
    # Iterate the loop to read the cell values
    for row in dataframe1.iter_rows(min_row=2, max_row=dataframe1.max_row, min_col=1, max_col=1):
        if row[0].value == part_nb:
            matched_row = row[0].row
            break

    if matched_row:
        part_nm = dataframe1.cell(row=matched_row, column=2).value
        clp = dataframe1.cell(row=matched_row, column=3).value
        cycle_lm = int(dataframe1.cell(row=matched_row, column=4).value)
        pro_rt = (clp / cycle_lm) * cycle_rm
        
        #print the results
        print("\n---Results---\n") 
        print("Part Name: " + part_nm)
        print(f"CLP: ${clp}")
        # Loop for printing different pro rates
        for percent in range(100, 0, -5):
            print(f"{percent}% Pro Rate: ${pro_rt * (percent / 100):.2f}")

    else:
        print('The inputted Part Number does not match any parts in the current database.')
    
    if input("Continue? (Y/N): ").upper() == 'N':
        print("Ending Program...")
        break